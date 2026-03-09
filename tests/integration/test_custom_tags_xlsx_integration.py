from __future__ import annotations

import importlib
import io
import unittest

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from services.common import db as dbm
from tests._helpers import basic_auth_header, temp_env


class TestCustomTagsXlsxIntegration(unittest.TestCase):
    def _new_client(self):
        mod = importlib.import_module("services.factory_api.app")
        importlib.reload(mod)
        return mod, TestClient(mod.app)

    def _seed_data(self, env) -> None:
        conn = dbm.connect(env)
        try:
            dbm.migrate(conn)
            conn.execute(
                """
                INSERT INTO channels(slug, display_name, kind, weight, render_profile, autopublish_enabled)
                VALUES(?, ?, 'LONG', 1.0, 'long_1080p24', 0)
                """,
                ("darkwood-reverie", "Darkwood Reverie"),
            )
            conn.execute("INSERT INTO canon_channels(value) VALUES(?)", ("darkwood-reverie",))
            conn.execute("INSERT INTO canon_thresholds(value) VALUES(?)", ("darkwood-reverie",))

            track_pk = int(
                conn.execute(
                    """
                    INSERT INTO tracks(channel_slug, track_id, gdrive_file_id, source, filename, title, artist, duration_sec, discovered_at, analyzed_at)
                    VALUES(?, ?, ?, 'gdrive', ?, ?, ?, 180.0, 1000.0, 1005.0)
                    """,
                    ("darkwood-reverie", "001", "file-001", "001.wav", "Title 001", "Artist X"),
                ).lastrowid
            )

            entries = [
                ("mist", "Mist", "VISUAL", "MANUAL"),
                ("aurora", "Aurora", "VISUAL", "AUTO"),
                ("serene", "Serene", "MOOD", "AUTO"),
                ("dream", "Dream", "THEME", "AUTO"),
            ]
            for code, label, category, state in entries:
                tag_id = int(
                    conn.execute(
                        """
                        INSERT INTO custom_tags(code, label, category, description, is_active, created_at, updated_at)
                        VALUES(?, ?, ?, '', 1, ?, ?)
                        """,
                        (code, label, category, "2024-01-01T00:00:00Z", "2024-01-01T00:00:00Z"),
                    ).lastrowid
                )
                conn.execute(
                    """
                    INSERT INTO track_custom_tag_assignments(track_pk, tag_id, state, assigned_at, updated_at)
                    VALUES(?, ?, ?, ?, ?)
                    """,
                    (track_pk, tag_id, state, "2024-01-01T00:00:00Z", "2024-01-01T00:00:00Z"),
                )
        finally:
            conn.close()

    def test_xlsx_custom_tag_columns_match_report_values(self) -> None:
        with temp_env() as (_, env):
            self._seed_data(env)
            _, client = self._new_client()
            auth = basic_auth_header(env.basic_user, env.basic_pass)

            api_resp = client.get(
                "/v1/track-catalog/analysis-report",
                params={"channel_slug": "darkwood-reverie"},
                headers=auth,
            )
            self.assertEqual(api_resp.status_code, 200)
            api_payload = api_resp.json()

            xlsx_resp = client.get(
                "/v1/track-catalog/analysis-report.xlsx",
                params={"channel_slug": "darkwood-reverie"},
                headers=auth,
            )
            self.assertEqual(xlsx_resp.status_code, 200)

            wb = load_workbook(io.BytesIO(xlsx_resp.content))
            ws = wb.active
            ordered_keys = [col["key"] for col in api_payload["columns"]]
            key_to_col = {key: idx for idx, key in enumerate(ordered_keys, start=1)}

            for key in ("custom_tags_visual", "custom_tags_mood", "custom_tags_theme"):
                self.assertIn(key, key_to_col)
                self.assertEqual(ws.cell(row=3, column=key_to_col[key]).value, api_payload["rows"][0][key])


if __name__ == "__main__":
    unittest.main()
