from __future__ import annotations

import importlib
import io
import json
import unittest

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from services.common import db as dbm
from services.track_analysis_report.xlsx_export import build_group_header_spans, sanitize_sheet_name
from tests._helpers import basic_auth_header, temp_env


class TestTrackAnalysisReportXlsxApi(unittest.TestCase):
    def _new_client(self):
        mod = importlib.import_module("services.factory_api.app")
        importlib.reload(mod)
        return mod, TestClient(mod.app)

    def _seed_channel(self, env) -> None:
        conn = dbm.connect(env)
        try:
            dbm.migrate(conn)
            conn.execute(
                """
                INSERT INTO channels(slug, display_name, kind, weight, render_profile, autopublish_enabled)
                VALUES(?, ?, 'LONG', 1.0, 'long_1080p24', 0)
                """,
                ("darkwood-reverie", "Darkwood Reverie / Night"),
            )
            conn.execute("INSERT INTO canon_channels(value) VALUES(?)", ("darkwood-reverie",))
            conn.execute("INSERT INTO canon_thresholds(value) VALUES(?)", ("darkwood-reverie",))
        finally:
            conn.close()

    def _seed_tracks(self, env) -> None:
        conn = dbm.connect(env)
        try:
            first_pk = int(
                conn.execute(
                    """
                    INSERT INTO tracks(channel_slug, track_id, gdrive_file_id, source, filename, title, artist, duration_sec, discovered_at, analyzed_at)
                    VALUES(?, ?, ?, 'gdrive', ?, ?, ?, 180.0, 1000.0, 1005.0)
                    """,
                    ("darkwood-reverie", "001", "file-001", "001.wav", "Title 001", "Artist X"),
                ).lastrowid
            )
            second_pk = int(
                conn.execute(
                    """
                    INSERT INTO tracks(channel_slug, track_id, gdrive_file_id, source, filename, title, artist, duration_sec, discovered_at, analyzed_at)
                    VALUES(?, ?, ?, 'gdrive', ?, ?, ?, 181.0, 2000.0, 2005.0)
                    """,
                    ("darkwood-reverie", "002", "file-002", "002.wav", "Title 002", "Artist Y"),
                ).lastrowid
            )
            conn.execute(
                "INSERT INTO track_features(track_pk, payload_json, computed_at) VALUES(?, ?, ?)",
                (
                    first_pk,
                    json.dumps(
                        {
                            "analysis_status": "ok",
                            "voice_flag": False,
                            "yamnet_agg": {"voice_labels_used": ["speech"], "speech_labels_used": ["conversation"]},
                            "advanced_v1": {
                                "meta": {"analyzer_version": "advanced_track_analyzer_v1.1", "schema_version": "advanced_v1"},
                                "quality": {"integrated_lufs": -18.2},
                                "dynamics": {"energy_mean": 0.44, "intensity_curve_summary": {"start_mean": 0.1}},
                                "timbre": {"brightness": 0.31},
                                "structure": {"intro_energy": 0.22, "section_summary": {"parts": 3}},
                                "voice": {"speech_probability": 0.12},
                                "similarity": {"normalized_feature_vector": [0.1, 0.2, 0.3], "diversity_penalty_base": 0.27},
                            },
                        }
                    ),
                    1010.0,
                ),
            )
            conn.execute(
                "INSERT INTO track_tags(track_pk, payload_json, computed_at) VALUES(?, ?, ?)",
                (
                    first_pk,
                    json.dumps(
                        {
                            "yamnet_tags": ["rain", "wind"],
                            "advanced_v1": {
                                "semantic": {"mood_tags": ["calm", "ambient"], "theme_tags": ["minimal"]},
                                "voice_tags": ["spoken_word"],
                                "classifier_evidence": {"yamnet_top_classes": [{"label": "rain", "score": 0.9}]},
                            },
                        }
                    ),
                    1020.0,
                ),
            )
            conn.execute(
                "INSERT INTO track_scores(track_pk, payload_json, computed_at) VALUES(?, ?, ?)",
                (
                    first_pk,
                    json.dumps(
                        {
                            "dsp_score": 0.93,
                            "advanced_v1": {
                                "semantic": {"functional_scores": {"focus": 0.8, "energy": 0.3, "narrative": 0.2, "background_compatibility": 0.7}},
                                "playlist_fit": {"continuity_score": 0.6, "mixability_score": 0.7, "variety_support_score": 0.8},
                                "transition": {"intro_profile": "soft", "outro_profile": "tail", "transition_risk_score": 0.1},
                                "suitability": {
                                    "content_type_fit_score": 0.9,
                                    "channel_fit_score": 0.85,
                                    "selected_content_context": "LONG_INSTRUMENTAL_AMBIENT",
                                    "content_type_fit_by_context": {"LONG_INSTRUMENTAL_AMBIENT": 0.9, "LONG_LYRICAL": 0.2},
                                },
                                "rule_trace": [{"rule_id": "semantic.focus.v1", "matched": True}],
                                "final_decisions": {"hard_veto": False, "soft_penalty_total": 0.15, "warning_codes": ["PENALTY_TRANSITION_RISK"]},
                            },
                        }
                    ),
                    1030.0,
                ),
            )
            conn.execute(
                """
                INSERT INTO track_analysis_flat(
                    track_pk, channel_slug, track_id, gdrive_file_id, analysis_computed_at,
                    analysis_status, analyzer_version, schema_version, duration_sec,
                    true_peak_dbfs, spikes_found, yamnet_top_tags_text, yamnet_top_classes_json,
                    voice_flag, voice_flag_reason, speech_flag, speech_flag_reason,
                    dominant_texture, texture_confidence, texture_reason, prohibited_cues_summary,
                    prohibited_cues_flags_json, dsp_score, dsp_score_version, dsp_notes,
                    legacy_scene, legacy_mood, legacy_safety, legacy_scene_match,
                    human_readable_notes, updated_at
                ) VALUES(
                    ?, ?, ?, ?, ?,
                    ?, ?, ?, ?,
                    ?, ?, ?, ?,
                    ?, ?, ?, ?,
                    ?, ?, ?, ?,
                    ?, ?, ?, ?,
                    ?, ?, ?, ?,
                    ?, ?
                )
                """,
                (
                    first_pk,
                    "darkwood-reverie",
                    "001",
                    "file-001",
                    1040.0,
                    "flat-ok",
                    "flat-analyzer-version",
                    "flat-schema-version",
                    180.0,
                    -0.1,
                    1,
                    "flat-rain, flat-wind",
                    "[]",
                    1,
                    "flat-voice-reason",
                    1,
                    "flat-speech-reason",
                    "flat-texture",
                    0.88,
                    "flat-texture-reason",
                    None,
                    "{}",
                    0.99,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "2026-01-01T00:00:00Z",
                ),
            )
            conn.execute(
                "INSERT INTO track_features(track_pk, payload_json, computed_at) VALUES(?, ?, ?)",
                (second_pk, '{"analysis_status":"ok","voice_flag":true}', 2010.0),
            )
        finally:
            conn.close()

    def test_xlsx_endpoint_matches_api_dataset(self) -> None:
        with temp_env() as (_, env):
            self._seed_channel(env)
            self._seed_tracks(env)
            _, client = self._new_client()
            auth = basic_auth_header(env.basic_user, env.basic_pass)

            api_resp = client.get(
                "/v1/track-catalog/analysis-report",
                params={"channel_slug": "darkwood-reverie"},
                headers=auth,
            )
            self.assertEqual(api_resp.status_code, 200)
            api_payload = api_resp.json()

            xlsx_resp = client.get(
                "/v1/track-catalog/analysis-report.xlsx",
                params={"channel_slug": "darkwood-reverie"},
                headers=auth,
            )

            self.assertEqual(xlsx_resp.status_code, 200)
            self.assertEqual(
                xlsx_resp.headers.get("content-type"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            disposition = xlsx_resp.headers.get("content-disposition") or ""
            self.assertRegex(
                disposition,
                r'^attachment; filename="analysis_report__darkwood-reverie__\d{8}_\d{6}\.xlsx"$',
            )

            wb = load_workbook(io.BytesIO(xlsx_resp.content))
            self.assertEqual(len(wb.sheetnames), 1)
            ws = wb.active

            expected_sheet = sanitize_sheet_name("Darkwood Reverie / Night")
            self.assertEqual(ws.title, expected_sheet)
            self.assertEqual(ws.freeze_panes, "A3")

            columns = api_payload["columns"]
            expected_spans = build_group_header_spans(columns)
            actual_merged = sorted(str(item) for item in ws.merged_cells.ranges)
            expected_merged = sorted(
                f"{_col(start)}1:{_col(end)}1" for start, end, _ in expected_spans
            )
            self.assertEqual(actual_merged, expected_merged)

            for start, _end, group in expected_spans:
                self.assertEqual(ws.cell(row=1, column=start).value, group)

            ordered_keys = [col["key"] for col in columns]
            self.assertEqual(
                [ws.cell(row=2, column=idx).value for idx in range(1, len(ordered_keys) + 1)],
                ordered_keys,
            )

            self.assertEqual(len(api_payload["rows"]), ws.max_row - 2)

            for row_offset, api_row in enumerate(api_payload["rows"], start=3):
                for col_idx, key in enumerate(ordered_keys, start=1):
                    self.assertEqual(ws.cell(row=row_offset, column=col_idx).value, api_row.get(key))

    def test_xlsx_includes_selected_advanced_columns_and_values(self) -> None:
        with temp_env() as (_, env):
            self._seed_channel(env)
            self._seed_tracks(env)
            _, client = self._new_client()
            auth = basic_auth_header(env.basic_user, env.basic_pass)
            api_resp = client.get("/v1/track-catalog/analysis-report", params={"channel_slug": "darkwood-reverie"}, headers=auth)
            self.assertEqual(api_resp.status_code, 200)
            payload = api_resp.json()
            columns = payload["columns"]
            ordered_keys = [col["key"] for col in columns]
            for required_key in (
                "analyzer_version",
                "schema_version",
                "hard_veto",
                "soft_penalty_total",
                "warning_codes_json",
                "mood_tags_csv",
                "theme_tags_csv",
                "intensity_curve_summary_json",
                "section_summary_json",
                "normalized_feature_vector_json",
                "similarity_diversity_penalty_base",
                "rule_trace_json",
            ):
                self.assertIn(required_key, ordered_keys)

            row = payload["rows"][0]
            expected_row_fragment = {
                "analysis_status": "flat-ok",
                "voice_flag": True,
                "yamnet_tags": "flat-rain, flat-wind",
                "dominant_texture": "flat-texture",
                "dsp_score": 0.99,
                "analyzer_version": "flat-analyzer-version",
                "schema_version": "flat-schema-version",
                "hard_veto": False,
                "soft_penalty_total": 0.15,
                "mood_tags_csv": "calm, ambient",
                "theme_tags_csv": "minimal",
                "normalized_feature_vector_json": "[0.1, 0.2, 0.3]",
                "rule_trace_json": '[{"matched": true, "rule_id": "semantic.focus.v1"}]',
                "similarity_diversity_penalty_base": 0.27,
            }
            self.assertEqual({key: row[key] for key in expected_row_fragment}, expected_row_fragment)
            self.assertEqual(row["analyzer_version"], "flat-analyzer-version")
            self.assertEqual(row["schema_version"], "flat-schema-version")
            self.assertEqual(row["hard_veto"], False)
            self.assertEqual(row["soft_penalty_total"], 0.15)
            self.assertEqual(row["mood_tags_csv"], "calm, ambient")
            self.assertEqual(row["theme_tags_csv"], "minimal")
            self.assertEqual(row["similarity_diversity_penalty_base"], 0.27)

            xlsx_resp = client.get(
                "/v1/track-catalog/analysis-report.xlsx",
                params={"channel_slug": "darkwood-reverie"},
                headers=auth,
            )
            self.assertEqual(xlsx_resp.status_code, 200)
            wb = load_workbook(io.BytesIO(xlsx_resp.content))
            ws = wb.active
            header_to_col = {ws.cell(row=2, column=idx).value: idx for idx in range(1, ws.max_column + 1)}
            self.assertIn("analysis_status", header_to_col)
            self.assertIn("similarity_diversity_penalty_base", header_to_col)
            self.assertEqual(
                ws.cell(row=3, column=header_to_col["analysis_status"]).value,
                row["analysis_status"],
            )
            self.assertEqual(
                ws.cell(row=3, column=header_to_col["similarity_diversity_penalty_base"]).value,
                row["similarity_diversity_penalty_base"],
            )
            self.assertEqual(
                ws.cell(row=3, column=header_to_col["hard_veto"]).value,
                row["hard_veto"],
            )

    def test_xlsx_legacy_rows_without_advanced_v1_remain_valid(self) -> None:
        with temp_env() as (_, env):
            self._seed_channel(env)
            self._seed_tracks(env)
            _, client = self._new_client()
            auth = basic_auth_header(env.basic_user, env.basic_pass)

            xlsx_resp = client.get(
                "/v1/track-catalog/analysis-report.xlsx",
                params={"channel_slug": "darkwood-reverie"},
                headers=auth,
            )
            self.assertEqual(xlsx_resp.status_code, 200)

            wb = load_workbook(io.BytesIO(xlsx_resp.content))
            ws = wb.active
            header_to_col = {ws.cell(row=2, column=idx).value: idx for idx in range(1, ws.max_column + 1)}
            self.assertEqual(ws.max_row - 2, 2)
            self.assertEqual(ws.cell(row=3, column=header_to_col["analysis_status"]).value, "flat-ok")
            self.assertEqual(ws.cell(row=3, column=header_to_col["voice_flag"]).value, True)
            self.assertEqual(ws.cell(row=3, column=header_to_col["yamnet_tags"]).value, "flat-rain, flat-wind")
            self.assertEqual(ws.cell(row=3, column=header_to_col["dsp_score"]).value, 0.99)
            self.assertEqual(ws.cell(row=3, column=header_to_col["analyzer_version"]).value, "flat-analyzer-version")
            self.assertEqual(ws.cell(row=3, column=header_to_col["soft_penalty_total"]).value, 0.15)



def _col(value: int) -> str:
    result = ""
    current = value
    while current > 0:
        current, rem = divmod(current - 1, 26)
        result = chr(65 + rem) + result
    return result


if __name__ == "__main__":
    unittest.main()
