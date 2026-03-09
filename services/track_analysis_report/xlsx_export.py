from __future__ import annotations

import re
from collections.abc import Sequence
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

_INVALID_SHEET_CHARS_RE = re.compile(r"[\\/*?:\[\]]")
_MAX_SHEET_NAME_LENGTH = 31


def sanitize_sheet_name(value: str) -> str:
    cleaned = _INVALID_SHEET_CHARS_RE.sub("_", str(value or "")).strip()
    if not cleaned:
        cleaned = "Sheet1"
    return cleaned[:_MAX_SHEET_NAME_LENGTH]


def build_group_header_spans(columns: Sequence[dict[str, Any]]) -> list[tuple[int, int, str]]:
    spans: list[tuple[int, int, str]] = []
    if not columns:
        return spans

    start_idx = 1
    current_group = str(columns[0].get("group") or "")
    for idx, col in enumerate(columns[1:], start=2):
        next_group = str(col.get("group") or "")
        if next_group != current_group:
            spans.append((start_idx, idx - 1, current_group))
            start_idx = idx
            current_group = next_group
    spans.append((start_idx, len(columns), current_group))
    return spans


def export_report_to_xlsx_bytes(report: dict[str, Any], sheet_title: str) -> bytes:
    columns = list(report.get("columns") or [])
    rows = list(report.get("rows") or [])

    wb = Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_name(sheet_title)

    header_font = Font(bold=True)

    for start_col, end_col, group in build_group_header_spans(columns):
        ws.cell(row=1, column=start_col, value=group)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        group_cell = ws.cell(row=1, column=start_col)
        group_cell.font = header_font
        group_cell.alignment = Alignment(horizontal="center")

    ordered_keys = [str(col.get("key") or "") for col in columns]
    for col_idx, key in enumerate(ordered_keys, start=1):
        header_cell = ws.cell(row=2, column=col_idx, value=key)
        header_cell.font = header_font

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, key in enumerate(ordered_keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    ws.freeze_panes = "A3"
    if ordered_keys:
        end_col_letter = get_column_letter(len(ordered_keys))
        ws.auto_filter.ref = f"A2:{end_col_letter}{max(ws.max_row, 2)}"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
